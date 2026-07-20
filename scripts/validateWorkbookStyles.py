"""
Character.xlsx の書式・構造の検証（Phase 0.75）。

  python scripts/validateWorkbookStyles.py

SheetJS(Community) は書式を保存できないため、**Character.xlsx を書き換える処理は openpyxl を使うこと**。
このスクリプトは、SheetJS で誤って書き戻された場合に失われる要素（列幅・太字・罫線・塗り・結合）を検出する。
問題があれば非0終了する。
"""
import io
import os
import sys
import unicodedata

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

XLSX = os.path.join("assets", "characters", "Character.xlsx")
WORLD_SHEETS = ["ground", "waterside", "sky", "bug", "phantom", "planet"]
OFFICIAL_SHEETS = WORLD_SHEETS + ["master_prompt", "unresolved"]
VALID_RARITIES = {"normal", "rare", "legendary", "secret"}
WHITE_TIGER_ID = "ground_rare_white_tiger"


def nk(v):
    return "" if v is None else unicodedata.normalize("NFKC", str(v)).strip()


def col_of(ws, header):
    for c in ws[1]:
        if nk(c.value).lower() == header.lower():
            return c.column
    return -1


def count_style(ws, limit=400):
    bold = borders = fills = 0
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, limit)):
        for c in row:
            if c.font and c.font.bold:
                bold += 1
            if c.border and any(getattr(c.border, s) and getattr(c.border, s).style for s in ("left", "right", "top", "bottom")):
                borders += 1
            if c.fill and c.fill.fill_type:
                fills += 1
    return bold, borders, fills


def main():
    errors = []
    wb = openpyxl.load_workbook(XLSX)

    # 正式8シートは先頭に順序どおり必須。制作用シート（reference_catalog / generation_queue /
    # real_master_prompt など）の追記は許容する。export:master はホワイトリスト方式なので取り込まれない。
    if wb.sheetnames[: len(OFFICIAL_SHEETS)] != OFFICIAL_SHEETS:
        errors.append(
            f"正式8シートが先頭に順序どおり存在しない: {wb.sheetnames[: len(OFFICIAL_SHEETS)]}（期待 {OFFICIAL_SHEETS}）"
        )

    ids = []
    for w in WORLD_SHEETS:
        if w not in wb.sheetnames:
            errors.append(f"シート欠落: {w}")
            continue
        ws = wb[w]

        # 書式（SheetJS で書き戻すと全て 0 になる）
        widths = [k for k, v in ws.column_dimensions.items() if v.width]
        bold, borders, fills = count_style(ws)
        if not widths:
            errors.append(f"{w}: 列幅が失われている（openpyxl 以外で保存した可能性）")
        if bold == 0:
            errors.append(f"{w}: 太字（ヘッダ書式）が失われている")
        if borders == 0:
            errors.append(f"{w}: 罫線が失われている")
        if fills == 0:
            errors.append(f"{w}: セルの塗りが失われている")

        # id 列
        i_id, i_en, i_rar = col_of(ws, "id"), col_of(ws, "英名"), col_of(ws, "rarity")
        if i_id < 0:
            errors.append(f"{w}: id 列が無い")
            continue
        if i_id != i_en + 1:
            errors.append(f"{w}: id 列が英名の直後にない（id={i_id}, 英名={i_en}）")
        if not ws.column_dimensions[ws.cell(1, i_id).column_letter].width:
            errors.append(f"{w}: id 列に列幅が設定されていない")
        if not (ws.cell(1, i_id).font and ws.cell(1, i_id).font.bold):
            errors.append(f"{w}: id ヘッダが他ヘッダと整合していない（太字でない）")

        for r in range(2, ws.max_row + 1):
            cid, en, rar = nk(ws.cell(r, i_id).value), nk(ws.cell(r, i_en).value), nk(ws.cell(r, i_rar).value).lower()
            if not cid and not en:
                continue
            if not cid:
                errors.append(f"{w} 行{r}: id 空欄")
            if not en:
                errors.append(f"{w} 行{r}: 英名 空欄")
            if rar == "legend":
                errors.append(f"{w} 行{r}: legend が残っている")
            elif rar not in VALID_RARITIES:
                errors.append(f"{w} 行{r}: rarity 不正 '{rar}'")
            if cid:
                ids.append(cid)
            # 文字列として保存され、指数表記・日付化していないこと
            if cid and not isinstance(ws.cell(r, i_id).value, str):
                errors.append(f"{w} 行{r}: id が文字列でない（{type(ws.cell(r, i_id).value).__name__}）")

    if len(ids) != len(set(ids)):
        errors.append(f"ID 重複あり（{len(ids)} 件中ユニーク {len(set(ids))} 件）")

    # White Tiger は rare
    ws = wb["ground"]
    i_id, i_rar = col_of(ws, "id"), col_of(ws, "rarity")
    found = False
    for r in range(2, ws.max_row + 1):
        if nk(ws.cell(r, i_id).value) == WHITE_TIGER_ID:
            found = True
            if nk(ws.cell(r, i_rar).value).lower() != "rare":
                errors.append(f"White Tiger の rarity が rare でない: {ws.cell(r, i_rar).value}")
    if not found:
        errors.append(f"{WHITE_TIGER_ID} が Excel に無い")

    # unresolved（管理列と麒麟行）
    if "unresolved" in wb.sheetnames:
        wsu = wb["unresolved"]
        hdr = [nk(c.value) for c in wsu[1]]
        for need in ("originalSheet", "originalExcelRow", "unresolvedReason"):
            if need not in hdr:
                errors.append(f"unresolved: {need} 列が無い")
        if wsu.max_row < 2:
            errors.append("unresolved: 退避行が無い")

    # master_prompt
    if "master_prompt" in wb.sheetnames:
        wsm = wb["master_prompt"]
        if wsm.max_row < 1 or not any(nk(c.value) for row in wsm.iter_rows() for c in row):
            errors.append("master_prompt の内容が失われている")

    print("=== Character.xlsx 書式・構造検証 ===")
    print(f"  シート: {wb.sheetnames}")
    print(f"  正式6シートの ID 件数: {len(ids)}")
    if errors:
        print(f"\nFAILED: {len(errors)} 件")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)
    print("\nOK: 書式・構造・ID・rarity すべて正常。")


if __name__ == "__main__":
    main()
