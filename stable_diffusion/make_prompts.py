from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
DEFAULT_DOCS_DIR = REPO_ROOT / "docs"
DEFAULT_PROMPTS_PATH = SCRIPT_DIR / "prompts_test_10.txt"
DEFAULT_ORDER_PATH = SCRIPT_DIR / "generation_order_test_10.csv"


@dataclass(frozen=True)
class CharacterRow:
    sheet_name: str
    row_no: int
    character_name: str
    motif: str
    features: str


def normalize_cell(value: object) -> str:
    return str(value).strip() if value is not None else ""


def find_default_input() -> Path:
    candidates = [
        path
        for path in DEFAULT_DOCS_DIR.glob("*.xlsx")
        if not path.name.startswith("~$")
    ]
    if candidates:
        return sorted(candidates)[0]

    csv_candidates = [path for path in DEFAULT_DOCS_DIR.glob("*.csv")]
    if csv_candidates:
        return sorted(csv_candidates)[0]

    raise FileNotFoundError(
        "docsフォルダにExcelまたはCSVが見つかりません。--input で入力ファイルを指定してください。"
    )


def read_csv_rows(path: Path) -> list[CharacterRow]:
    encodings = ("utf-8-sig", "utf-8", "cp932")
    last_error: Exception | None = None

    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                reader = csv.DictReader(file)
                rows: list[CharacterRow] = []
                for index, row in enumerate(reader, start=2):
                    character_name = normalize_cell(
                        row.get("キャラ名") or row.get("character_name") or row.get("name")
                    )
                    motif = normalize_cell(row.get("モチーフ") or row.get("motif"))
                    features = normalize_cell(row.get("特徴") or row.get("features"))

                    if not character_name or not motif:
                        continue

                    rows.append(
                        CharacterRow(
                            sheet_name=path.stem,
                            row_no=index,
                            character_name=character_name,
                            motif=motif,
                            features=features,
                        )
                    )
                return rows
        except UnicodeDecodeError as error:
            last_error = error

    raise UnicodeDecodeError(
        "csv",
        b"",
        0,
        1,
        f"CSVの文字コードを読み取れませんでした: {last_error}",
    )


def read_excel_rows(path: Path, sheet_name: str | None) -> list[CharacterRow]:
    try:
        import openpyxl
    except ImportError as error:
        raise RuntimeError(
            "Excelを読むには openpyxl が必要です。`pip install openpyxl` を実行するか、CSVを指定してください。"
        ) from error

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    worksheets = [workbook[sheet_name]] if sheet_name else workbook.worksheets
    rows: list[CharacterRow] = []

    for worksheet in worksheets:
        header_row_no: int | None = None
        header_map: dict[str, int] = {}

        for row_no, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            values = [normalize_cell(value) for value in row]
            if "No" in values and "キャラ名" in values:
                header_row_no = row_no
                header_map = {value: index for index, value in enumerate(values) if value}
                break

        if header_row_no is None:
            continue

        name_index = header_map["キャラ名"]
        motif_index = header_map.get("モチーフ")
        features_index = header_map.get("特徴")

        for row_no, row in enumerate(
            worksheet.iter_rows(min_row=header_row_no + 1, values_only=True),
            start=header_row_no + 1,
        ):
            cells = list(row)
            character_name = normalize_cell(cells[name_index] if name_index < len(cells) else "")
            motif = normalize_cell(cells[motif_index] if motif_index is not None and motif_index < len(cells) else "")
            features = normalize_cell(
                cells[features_index] if features_index is not None and features_index < len(cells) else ""
            )

            if not character_name or not motif:
                continue

            rows.append(
                CharacterRow(
                    sheet_name=worksheet.title,
                    row_no=row_no,
                    character_name=character_name,
                    motif=motif,
                    features=features,
                )
            )

    return rows


def read_character_rows(path: Path, sheet_name: str | None) -> list[CharacterRow]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_excel_rows(path, sheet_name)
    raise ValueError(f"対応していない入力形式です: {path.suffix}")


def build_prompt(row: CharacterRow) -> str:
    parts = [
        "WORLDAWN original cute 2D mascot creature",
        "single full-body character",
        "one creature only",
        "centered composition",
        "clean white background",
        "simple readable silhouette",
        "smartphone collection game character",
        "fantasy creature design",
        "clean line art",
        "soft cel shading",
        "high quality",
        "no text",
        "no logo",
        "no UI",
        f"character name concept: {row.character_name}",
        f"motif: {row.motif}",
        f"key design feature: {row.features}",
        f"make the {row.motif} motif easy to recognize",
        "designed for a creature encyclopedia card",
    ]
    return ", ".join(part for part in parts if part)


def limit_rows(rows: list[CharacterRow], limit: int | None) -> list[CharacterRow]:
    return rows if limit is None else rows[:limit]


def write_prompts(rows: Iterable[CharacterRow], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    prompts = [build_prompt(row) for row in rows]
    path.write_text("\n".join(prompts) + "\n", encoding="utf-8-sig")


def write_generation_order(rows: Iterable[CharacterRow], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "generation_no",
                "sheet_name",
                "row_no",
                "character_name",
                "motif",
                "features",
                "prompt",
            ],
        )
        writer.writeheader()
        for generation_no, row in enumerate(rows, start=1):
            writer.writerow(
                {
                    "generation_no": generation_no,
                    "sheet_name": row.sheet_name,
                    "row_no": row.row_no,
                    "character_name": row.character_name,
                    "motif": row.motif,
                    "features": row.features,
                    "prompt": build_prompt(row),
                }
            )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="WORLDAWN管理表からStable Diffusion WebUI Forge用プロンプトを作成します。"
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=None,
        help="入力Excel/CSV。省略時は docs フォルダ内のExcel/CSVを自動検出します。",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Excelのシート名。省略時は全シートを順番に読みます。",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=10,
        help="出力件数。初期値は10件です。",
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="全件出力します。--limit より優先されます。",
    )
    parser.add_argument(
        "--out-prompts",
        type=Path,
        default=DEFAULT_PROMPTS_PATH,
        help="Forgeに貼り付けるプロンプト一覧の出力先。",
    )
    parser.add_argument(
        "--out-order",
        type=Path,
        default=DEFAULT_ORDER_PATH,
        help="生成順対応CSVの出力先。",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = args.input.resolve() if args.input else find_default_input()
    rows = read_character_rows(input_path, args.sheet)
    selected_rows = limit_rows(rows, None if args.all else args.limit)

    if not selected_rows:
        raise RuntimeError("出力対象のキャラクターが見つかりませんでした。入力ファイルを確認してください。")

    write_prompts(selected_rows, args.out_prompts)
    write_generation_order(selected_rows, args.out_order)

    print(f"入力: {input_path}")
    print(f"出力件数: {len(selected_rows)}")
    print(f"プロンプト: {args.out_prompts}")
    print(f"生成順CSV: {args.out_order}")


if __name__ == "__main__":
    main()
